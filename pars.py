from openpyxl import load_workbook

print('Введите название файла прайс-листа магазина i-car(название должно быть строго указано с расширением)')
name_of_icar_price_list = input()
print('Введите название листа из файла магазина i-car(https://imgur.com/a/8NEkt9W)')
name_of_icar_price_list_sheet = input()

print('Введите название файла прайс-листа дистрибьютора(название должно быть строго указано с расширением)')
name_of_distributor_price_list = input()
print('Введите название листа из файла магазина дистрибьютора(https://imgur.com/a/8NEkt9W)')
name_of_distributor_price_list_sheet = input()

article_name_dict_icar = {}
article_name_dict_dist = {}

f = open("C:/programs for work/not in icar.txt", 'w', encoding='utf-8')

wb1 = load_workbook(name_of_icar_price_list)
ws1 = wb1[name_of_icar_price_list_sheet]

wb2 = load_workbook(name_of_distributor_price_list)
ws2 = wb2[name_of_distributor_price_list_sheet]

article_icar = ws1['A']
article_dist = ws2['A']

name_icar = ws1['B']
name_dist = ws2['B']

column_list1 = [article_icar[x].value for x in range(len(article_icar))]
column_list2 = [article_dist[y].value for y in range(len(article_dist))]

column2_list1 = [name_icar[x].value for x in range(len(name_icar))]
column2_list2 = [name_dist[y].value for y in range(len(name_dist))]

column2_list1.pop(0)
column2_list2.pop(0)


i = 1
while i < len(column_list1):
    if column_list1[i] == '':
        column_list1.pop(i)
        continue
    column_list1[i] = int(str(column_list1[i]))
    i += 1

k = 1
while k < len(column_list2):
    if column_list2[k] == '':
        column_list2.pop(k)
        continue
    column_list2[k] = int(str(column_list2[k]))
    k += 1

column_list1.pop(0)
column_list2.pop(0)

for temp1 in range(len(column_list1)):
    article_name_dict_icar[column_list1[temp1]] = column2_list1[temp1]
for temp2 in range(len(column_list2)):
    article_name_dict_dist[column_list2[temp2]] = column2_list2[temp2]

not_in = []
for k in range(len(column_list2)):
    if column_list2[k] not in column_list1:
        column_list2[k] = int(column_list2[k])
        not_in.append(column_list2[k])

print(len(not_in))

for x in range(len(not_in)):
    f.write(str(not_in[x]))
    f.write(', ')
    f.write(article_name_dict_dist[not_in[x]])
    f.write('\n')

f.close()

input()